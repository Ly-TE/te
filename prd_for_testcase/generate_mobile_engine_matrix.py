from pathlib import Path
import zipfile
import re
from datetime import datetime
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DOCX_PATH = Path(r"prd_for_testcase/手游引擎测试.docx")
OUTPUT_PATH = Path(r"prd_for_testcase/手游引擎测试结果矩阵.xlsx")


def extract_docx_text(docx_path: Path) -> str:
    with zipfile.ZipFile(docx_path, "r") as zf:
        xml_content = zf.read("word/document.xml")
    root = ET.fromstring(xml_content)
    namespaces = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for p in root.findall(".//w:p", namespaces):
        texts = [t.text or "" for t in p.findall(".//w:t", namespaces)]
        paragraph = "".join(texts).strip()
        if paragraph:
            paragraphs.append(paragraph)
    return "\n".join(paragraphs)


def extract_case_titles(text: str) -> list[str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    stop_markers = ["王征", "Json模型"]
    titles = []
    for line in lines:
        if any(marker in line for marker in stop_markers):
            break
        if re.search(r"\d{1,2}/\d{1,2}", line):
            continue
        if len(line) > 60:
            continue
        titles.append(line)
    # 去重并保持顺序
    seen = set()
    unique_titles = []
    for title in titles:
        if title not in seen:
            seen.add(title)
            unique_titles.append(title)
    return unique_titles


def expand_titles_to_columns(titles: list[str]) -> list[str]:
    """将第二行内容按顿号/逗号拆分为独立列。"""
    expanded = []
    for title in titles:
        parts = [part.strip() for part in re.split(r"[、，,]", title) if part.strip()]
        if len(parts) > 1:
            expanded.extend(parts)
        else:
            expanded.append(title)
    return expanded


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 24)


def build_workbook(case_titles: list[str], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试矩阵"

    ws.cell(row=1, column=1, value="手游安卓")
    ws.cell(row=1, column=2, value="手游IOS")

    for row_index, title in enumerate(case_titles, start=2):
        ws.cell(row=row_index, column=1, value=title)
        ws.cell(row=row_index, column=2, value=title)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        alt_path = output_path.with_name(
            f"{output_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
        )
        wb.save(alt_path)
        return alt_path


def main() -> None:
    text = extract_docx_text(DOCX_PATH)
    case_titles = extract_case_titles(text)
    expanded_titles = expand_titles_to_columns(case_titles)
    saved_path = build_workbook(expanded_titles, OUTPUT_PATH)
    print(f"Generated: {saved_path}")
    print(f"Case count: {len(expanded_titles)}")


if __name__ == "__main__":
    main()