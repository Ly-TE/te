from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUT = r"prd_for_testcase\xiaomi_埋点测试用例.xlsx"


def autosize_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def apply_style(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def build_workbook(output_path=OUT):
    headers = ["用例编号", "模块", "优先级", "用例标题", "前置条件", "测试步骤", "测试数据", "预期结果", "备注"]

    cases = [
        ["XM-MD-TC-001", "页面曝光-page_exposure", "P0", "验证进入游戏开始加速页面上报page_exposure事件", "埋点日志/抓包工具可用；用户可进入游戏开始加速页面", "1.清理或标记当前埋点日志\n2.进入游戏开始加速页面\n3.查看客户端埋点日志或服务端接收数据", "cur_page_category=game_acceleration_leishen\npage_package_name=当前页面App包名", "成功上报page_exposure事件；cur_page_category为game_acceleration_leishen；page_package_name为非空字符串且符合包名格式；事件无重复异常上报", "页面曝光核心场景"],
        ["XM-MD-TC-002", "页面曝光-page_exposure", "P0", "验证进入付费/订阅页面上报page_exposure事件", "埋点日志/抓包工具可用；用户可进入付费/订阅页面", "1.从加速页或侧栏进入付费/订阅页面\n2.查看埋点上报数据", "cur_page_category=recharge_leishen\npage_package_name=当前页面App包名", "成功上报page_exposure事件；cur_page_category为recharge_leishen；page_package_name正确；上报时机为页面展示时", "付费页曝光"],
        ["XM-MD-TC-003", "页面曝光-page_exposure", "P0", "验证进入加速结果页面上报page_exposure事件", "用户已完成一次加速或可进入加速结果页面", "1.触发加速完成/查看结果\n2.进入加速结果页面\n3.查看埋点上报数据", "cur_page_category=acceleration_results_leishen\npage_package_name=当前页面App包名", "成功上报page_exposure事件；cur_page_category为acceleration_results_leishen；page_package_name正确", "结果页曝光"],
        ["XM-MD-TC-004", "页面曝光-page_exposure", "P1", "验证进入订单页面上报page_exposure事件", "用户可进入订单页面；埋点日志/抓包工具可用", "1.进入订单页面\n2.查看埋点上报数据", "cur_page_category=order_leishen\npage_package_name=当前页面App包名", "成功上报page_exposure事件；cur_page_category为order_leishen；page_package_name正确", "订单页曝光"],
        ["XM-MD-TC-005", "页面曝光-page_exposure", "P1", "验证页面曝光事件字段类型和枚举值合法", "已覆盖进入加速、付费、结果、订单页面", "1.分别进入PRD列出的4类页面\n2.导出page_exposure事件\n3.检查字段类型和枚举", "cur_page_category枚举：game_acceleration_leishen/recharge_leishen/acceleration_results_leishen/order_leishen\npage_package_name=str", "page_exposure仅使用PRD定义枚举；字段均为str类型；不存在空值、错别字、中英文冒号混用导致的异常值", "字段规范校验"],

        ["XM-MD-TC-006", "元素曝光-item_exposure", "P0", "验证开始加速按钮曝光上报item_exposure事件", "用户进入包含开始加速按钮的页面", "1.进入游戏开始加速页面\n2.确保开始加速按钮展示在可视区域\n3.查看埋点上报数据", "item_type=button\nitem_enum=start_accelerating_leishen\npackage_name=当前游戏包名", "成功上报item_exposure事件；item_type为button；item_enum为start_accelerating_leishen；package_name非空且正确", "按钮曝光"],
        ["XM-MD-TC-007", "元素曝光-item_exposure", "P0", "验证付费充值按钮曝光上报item_exposure事件", "用户进入包含付费充值按钮的页面", "1.进入付费/订阅页面或付费入口区域\n2.确保付费充值按钮展示\n3.查看埋点上报数据", "item_type=button\nitem_enum=pay_leishen\npackage_name=当前游戏包名", "成功上报item_exposure事件；item_enum为pay_leishen；package_name正确", "按钮曝光"],
        ["XM-MD-TC-008", "元素曝光-item_exposure", "P1", "验证游戏启动按钮曝光上报item_exposure事件", "用户进入包含游戏启动按钮的页面", "1.进入加速结果页或可启动游戏页面\n2.确保游戏启动按钮展示\n3.查看埋点上报数据", "item_type=button\nitem_enum=open_leishen\npackage_name=当前游戏包名", "成功上报item_exposure事件；item_enum为open_leishen；package_name正确", "按钮曝光"],
        ["XM-MD-TC-009", "元素曝光-item_exposure", "P1", "验证关闭加速按钮曝光上报item_exposure事件", "用户处于加速中且页面展示关闭加速按钮", "1.开启加速\n2.等待关闭加速按钮展示\n3.查看埋点上报数据", "item_type=button\nitem_enum=stop_accelerate_leishen\npackage_name=当前游戏包名", "成功上报item_exposure事件；item_enum为stop_accelerate_leishen；package_name正确", "按钮曝光"],
        ["XM-MD-TC-010", "元素曝光-item_exposure", "P0", "验证服务切换弹窗曝光上报item_exposure事件", "具备触发服务切换弹窗的条件", "1.执行触发服务切换弹窗的操作\n2.弹窗展示后查看埋点上报数据", "item_type=pop\nitem_enum=service_switch_leishen\npackage_name=当前游戏包名", "弹窗展示时成功上报item_exposure事件；item_type为pop；item_enum为service_switch_leishen；package_name正确", "弹窗曝光"],
        ["XM-MD-TC-011", "元素曝光-item_exposure", "P0", "验证时间限制弹窗曝光上报item_exposure事件", "具备触发时间限制弹窗的条件，如免费时长不足/用尽", "1.构造时间限制触发条件\n2.触发并展示时间限制弹窗\n3.查看埋点上报数据", "item_type=pop\nitem_enum=time_limit_leishen\npackage_name=当前游戏包名", "弹窗展示时成功上报item_exposure事件；item_type为pop；item_enum为time_limit_leishen；package_name正确", "弹窗曝光"],
        ["XM-MD-TC-012", "元素曝光-item_exposure", "P1", "验证元素重复曝光控制符合预期", "页面存在可重复进入或滚动展示的按钮/弹窗", "1.首次进入页面记录曝光\n2.停留页面不操作观察是否重复上报\n3.离开后重新进入页面再次观察", "任一button/pop曝光元素", "首次实际展示时上报曝光；同一次展示期间不出现高频重复上报；重新进入或重新展示时按产品埋点规则再次上报", "防重复上报"],
        ["XM-MD-TC-013", "元素曝光-item_exposure", "P1", "验证元素曝光事件字段类型和枚举值合法", "已触发PRD列出的按钮和弹窗曝光", "1.导出item_exposure事件\n2.检查item_type、item_enum、package_name字段", "item_type=button/pop\nitem_enum为PRD定义枚举\npackage_name=str", "字段类型均为str；button仅对应按钮枚举，pop仅对应弹窗枚举；不存在未知枚举、空值或字段缺失", "字段规范校验"],

        ["XM-MD-TC-014", "元素点击-item_click", "P0", "验证点击开始加速按钮上报item_click事件", "开始加速按钮已展示且可点击", "1.进入游戏开始加速页面\n2.点击开始加速按钮\n3.查看埋点上报数据", "item_type=button\nitem_enum=start_accelerating_leishen\npackage_name=当前游戏包名", "点击时成功上报item_click事件；item_type为button；item_enum为start_accelerating_leishen；package_name正确；上报时机为点击发生时", "按钮点击核心场景"],
        ["XM-MD-TC-015", "元素点击-item_click", "P0", "验证点击付费充值按钮上报item_click事件", "付费充值按钮已展示且可点击", "1.进入付费入口或订阅页面\n2.点击付费充值按钮\n3.查看埋点上报数据", "item_type=button\nitem_enum=pay_leishen\npackage_name=当前游戏包名", "点击时成功上报item_click事件；item_enum为pay_leishen；package_name正确", "按钮点击"],
        ["XM-MD-TC-016", "元素点击-item_click", "P1", "验证点击游戏启动按钮上报item_click事件", "游戏启动按钮已展示且可点击", "1.进入加速结果页或可启动游戏页面\n2.点击游戏启动按钮\n3.查看埋点上报数据", "item_type=button\nitem_enum=open_leishen\npackage_name=当前游戏包名", "点击时成功上报item_click事件；item_enum为open_leishen；package_name正确", "按钮点击"],
        ["XM-MD-TC-017", "元素点击-item_click", "P0", "验证点击关闭加速按钮上报item_click事件", "用户处于加速中，关闭加速按钮已展示", "1.开启加速\n2.点击关闭加速按钮\n3.查看埋点上报数据", "item_type=button\nitem_enum=stop_accelerate_leishen\npackage_name=当前游戏包名", "点击时成功上报item_click事件；item_enum为stop_accelerate_leishen；package_name正确", "按钮点击"],
        ["XM-MD-TC-018", "元素点击-item_click", "P1", "验证按钮点击事件不上报弹窗枚举", "已覆盖所有按钮点击场景", "1.分别点击开始加速、付费充值、游戏启动、关闭加速按钮\n2.导出item_click事件检查字段", "item_type=button\nitem_enum为4个按钮枚举之一", "item_click事件item_type固定为button；item_enum不出现service_switch_leishen、time_limit_leishen等pop枚举", "枚举边界"],
        ["XM-MD-TC-019", "元素点击-item_click", "P1", "验证快速连续点击按钮时点击埋点记录准确", "目标按钮可点击；埋点日志可区分时间戳", "1.对目标按钮快速连续点击2-3次\n2.查看业务是否防重\n3.核对item_click上报次数和时间", "任一按钮枚举", "埋点上报次数与实际有效点击/产品防重规则一致；不丢失有效点击；不产生与用户操作无关的额外点击事件", "异常/边界"],

        ["XM-MD-TC-020", "接口埋点-api_request", "P0", "验证发起加速请求时上报api_request事件", "用户具备加速资格；网络正常", "1.点击开始加速\n2.在请求发起时查看埋点上报数据", "api_path=accelerate_leishen\ntype=acceleration_request_leishen", "加速请求发起时成功上报api_request事件；api_path为accelerate_leishen；type为acceleration_request_leishen", "接口请求"],
        ["XM-MD-TC-021", "接口埋点-api_request", "P0", "验证加速成功时上报api_request成功事件", "可模拟或实际完成加速成功", "1.发起加速请求\n2.等待加速成功\n3.查看埋点上报数据", "api_path=accelerate_leishen\ntype=acceleration_success_leishen", "加速成功时成功上报api_request事件；type为acceleration_success_leishen；不误报失败类型", "接口结果"],
        ["XM-MD-TC-022", "接口埋点-api_request", "P0", "验证加速失败时上报api_request失败事件", "可模拟加速失败，如网络异常、服务异常或资格不足", "1.构造加速失败条件\n2.发起加速请求\n3.查看埋点上报数据", "api_path=accelerate_leishen\ntype=acceleration_failure_leishen", "加速失败时成功上报api_request事件；type为acceleration_failure_leishen；不误报成功类型", "接口异常"],
        ["XM-MD-TC-023", "接口埋点-api_request", "P1", "验证启动游戏时上报api_request事件", "游戏启动入口可用", "1.进入可启动游戏页面\n2.点击启动游戏\n3.查看埋点上报数据", "api_path=accelerate_leishen\ntype=start_game_leishen", "启动游戏时成功上报api_request事件；api_path为accelerate_leishen；type为start_game_leishen", "启动游戏"],
        ["XM-MD-TC-024", "接口埋点-api_request", "P1", "验证api_request事件字段类型和枚举值合法", "已触发加速请求、成功、失败、启动游戏场景", "1.导出api_request事件\n2.检查api_path和type字段", "api_path=accelerate_leishen\ntype为4个PRD定义枚举", "字段均为str类型；api_path固定为accelerate_leishen；type仅使用PRD定义枚举；不存在空值或未知值", "字段规范校验"],

        ["XM-MD-TC-025", "埋点通用校验", "P0", "验证一次完整加速成功链路埋点顺序和关键事件完整", "用户可完整完成进入页面、点击开始加速、加速成功、启动游戏流程", "1.进入游戏开始加速页面\n2.点击开始加速\n3.等待加速成功进入结果页\n4.点击启动游戏\n5.按时间顺序核对埋点", "page_exposure/item_exposure/item_click/api_request", "关键事件均有上报；大致顺序为页面曝光->按钮曝光->点击开始加速->加速请求->加速成功->结果页曝光/启动按钮曝光->点击启动游戏->启动游戏接口埋点", "链路完整性"],
        ["XM-MD-TC-026", "埋点通用校验", "P0", "验证一次加速失败链路埋点完整", "可构造加速失败条件", "1.进入游戏开始加速页面\n2.点击开始加速\n3.触发加速失败\n4.按时间顺序核对埋点", "item_click=start_accelerating_leishen\napi_request=acceleration_request_leishen/acceleration_failure_leishen", "点击、请求、失败事件均上报；失败场景不产生acceleration_success_leishen；错误页或弹窗如展示则按曝光规则上报", "失败链路"],
        ["XM-MD-TC-027", "埋点通用校验", "P1", "验证埋点字段不存在缺失、空值和类型错误", "已执行主要页面、按钮、弹窗、接口场景", "1.导出所有本需求相关埋点\n2.逐条检查必填字段、数据类型和枚举", "事件：page_exposure/item_exposure/item_click/api_request", "事件英文名正确；必填属性均存在；str字段未上传数字/对象/null；包名字段为实际包名或约定格式", "通用质量"],
        ["XM-MD-TC-028", "埋点通用校验", "P1", "验证不同游戏包名场景下package_name/page_package_name上报正确", "后台配置至少2个不同游戏或可切换测试包名", "1.分别进入A、B游戏加速链路\n2.触发页面曝光、元素曝光、点击和接口事件\n3.核对包名字段", "A包名、B包名", "不同游戏场景上报对应包名，不串包、不固定写死；page_package_name和package_name符合各事件定义", "多包名场景"],
        ["XM-MD-TC-029", "埋点通用校验", "P2", "验证弱网或离线恢复后埋点上报策略符合预期", "具备弱网/断网模拟能力；埋点SDK支持缓存或失败处理", "1.断网或弱网进入页面并操作按钮\n2.恢复网络\n3.查看埋点是否补发或按规则丢弃", "任一页面/按钮/接口埋点", "弱网场景不影响业务主流程；埋点按SDK策略补发或失败记录；不出现重复大量补发、字段损坏", "稳定性"],
        ["XM-MD-TC-030", "埋点通用校验", "P2", "验证埋点平台接收数据与客户端日志一致", "客户端日志和埋点平台查询权限可用", "1.执行一轮完整链路\n2.记录客户端本地上报日志\n3.在埋点平台按用户/设备/时间查询\n4.对比事件与字段", "同一测试账号/设备", "埋点平台可查询到对应事件；事件数量、时间、属性值与客户端上报日志一致或符合服务端处理规则", "端到端验收"],
    ]

    smoke_headers = ["冒烟编号", "关联用例编号", "冒烟场景", "验证重点"]
    smokes = [
        ["SMK-MD-001", "XM-MD-TC-001", "加速页面曝光", "进入游戏开始加速页面上报page_exposure且页面枚举正确"],
        ["SMK-MD-002", "XM-MD-TC-006", "开始加速按钮曝光", "按钮展示时上报item_exposure，item_type和item_enum正确"],
        ["SMK-MD-003", "XM-MD-TC-014", "开始加速按钮点击", "点击开始加速按钮上报item_click"],
        ["SMK-MD-004", "XM-MD-TC-020", "加速请求埋点", "发起加速时上报acceleration_request_leishen"],
        ["SMK-MD-005", "XM-MD-TC-021", "加速成功埋点", "加速成功时上报acceleration_success_leishen"],
        ["SMK-MD-006", "XM-MD-TC-022", "加速失败埋点", "加速失败时上报acceleration_failure_leishen"],
        ["SMK-MD-007", "XM-MD-TC-023", "启动游戏埋点", "点击启动游戏后上报start_game_leishen"],
        ["SMK-MD-008", "XM-MD-TC-027", "字段完整性", "所有事件必填字段存在、类型为str且枚举合法"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "埋点测试用例"
    ws.append(headers)
    for row in cases:
        ws.append(row)

    smoke = wb.create_sheet("冒烟测试用例")
    smoke.append(smoke_headers)
    for row in smokes:
        smoke.append(row)

    apply_style(ws)
    apply_style(smoke)
    autosize_columns(ws, [16, 22, 10, 42, 32, 52, 34, 52, 18])
    autosize_columns(smoke, [14, 18, 28, 48])

    wb.save(output_path)
    return output_path, len(cases), len(smokes)


if __name__ == "__main__":
    print(build_workbook())