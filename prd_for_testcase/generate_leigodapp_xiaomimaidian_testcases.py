from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUT = r"prd_for_testcase\leigodapp_xiaomi_埋点测试用例.xlsx"


def autosize_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def apply_style(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def build_workbook(output_path=OUT):
    headers = ["用例编号", "模块", "优先级", "用例标题", "前置条件", "测试步骤", "测试数据", "预期结果", "备注"]

    cases = [
        ["LG-XM-MD-TC-001", "事件映射", "P0", "验证小米APP_ACTIVE事件映射羚羊client_install", "测试包接入小米广告归因及羚羊埋点；可查看客户端日志、服务端日志或埋点平台数据", "1.通过小米广告渠道下载安装App\n2.打开App并完成登录\n3.查看小米埋点事件及羚羊事件上报数据", "小米事件=APP_ACTIVE\n羚羊事件=client_install", "登录成功后上报小米APP_ACTIVE；对应羚羊事件为client_install；事件名称、渠道标识、设备/用户标识正确", "PRD映射：APP_ACTIVE -> client_install"],
        ["LG-XM-MD-TC-002", "事件映射", "P0", "验证小米APP_REGISTER事件映射羚羊login_on", "测试设备为小米广告渠道首次安装且从未登录过；可查看埋点数据", "1.通过小米广告渠道首次下载安装App\n2.首次打开App并完成登录\n3.查看小米埋点事件及羚羊事件上报数据", "小米事件=APP_REGISTER\n羚羊事件=login_on", "设备首次登录成功后上报APP_REGISTER；对应羚羊事件为login_on；同一次首登链路数据可关联到正确设备与账号", "PRD映射：APP_REGISTER -> login_on"],
        ["LG-XM-MD-TC-003", "事件映射", "P0", "验证小米APP_PAY事件映射羚羊payment_successful", "用户通过付费广告下载App；账号可完成真实或沙箱支付；可查看支付成功回调和埋点数据", "1.通过付费广告渠道下载安装App\n2.登录账号并发起付费\n3.完成支付\n4.查看小米埋点事件及羚羊事件上报数据", "小米事件=APP_PAY\n羚羊事件=payment_successful", "支付成功后上报APP_PAY；对应羚羊事件为payment_successful；支付金额、订单号、用户/设备标识与订单系统一致（如有传参）", "PRD映射：APP_PAY -> payment_successful"],

        ["LG-XM-MD-TC-004", "APP_ACTIVE-激活", "P0", "验证小米广告渠道下载安装后登录成功上报APP_ACTIVE", "设备通过小米广告渠道下载安装App；账号可登录；埋点日志可用", "1.清理或标记当前埋点日志\n2.打开App\n3.完成手机号/账号登录\n4.查看登录成功后的埋点上报", "渠道=小米广告渠道\n登录结果=成功", "只要登录成功即上报APP_ACTIVE；上报时机在登录成功后；渠道来源识别为小米广告；羚羊client_install同步产生或可关联", "核心正向场景"],
        ["LG-XM-MD-TC-005", "APP_ACTIVE-激活", "P0", "验证老设备二次登录仍上报APP_ACTIVE", "同一设备曾通过小米广告渠道安装并登录过；当前已退出登录", "1.打开App\n2.使用同一账号或其他账号再次登录\n3.查看埋点上报", "设备=已登录过设备\n登录=二次登录成功", "二次登录成功后仍上报APP_ACTIVE；不因非首次登录而漏报激活事件；APP_REGISTER不应再次上报", "PRD说明：不限制是否首次登录"],
        ["LG-XM-MD-TC-006", "APP_ACTIVE-激活", "P1", "验证注册后重复登录上报APP_ACTIVE", "小米广告渠道设备；账号已完成注册和首次登录；可退出重新登录", "1.退出当前账号\n2.重新登录该账号\n3.查看埋点数据", "场景=注册后重复登录", "重复登录成功后上报APP_ACTIVE；不重复上报APP_REGISTER；事件用户标识为当前登录账号", "覆盖PRD举例场景"],
        ["LG-XM-MD-TC-007", "APP_ACTIVE-激活", "P1", "验证登录失败不应上报APP_ACTIVE", "通过小米广告渠道安装App；可构造登录失败，如密码错误/验证码错误/网络异常", "1.打开App\n2.输入错误登录信息或断网提交登录\n3.查看埋点上报", "登录结果=失败", "登录失败不产生APP_ACTIVE；不产生羚羊client_install；如有登录失败业务埋点应与本需求事件区分", "负向场景"],
        ["LG-XM-MD-TC-008", "APP_ACTIVE-激活", "P1", "验证非小米广告渠道安装登录不上报小米APP_ACTIVE", "设备通过自然量、其他广告渠道或本地安装包安装App；账号可登录", "1.从非小米广告渠道安装App\n2.完成登录\n3.查看小米埋点上报", "渠道=非小米广告渠道", "不向小米上报APP_ACTIVE；羚羊内部事件如按全量规则上报，应不被错误归因为小米广告渠道", "渠道边界"],

        ["LG-XM-MD-TC-009", "APP_REGISTER-注册/首登", "P0", "验证小米广告渠道首次打开并首次登录上报APP_REGISTER", "全新设备或已重置测试设备；通过小米广告渠道首次下载安装App；设备历史无登录记录", "1.首次打开App\n2.完成注册并登录，或使用新账号完成首次登录\n3.查看埋点数据", "设备首次登录=是", "设备第一次登录成功后上报APP_REGISTER一次；对应羚羊login_on；同时按PRD登录成功也应上报APP_ACTIVE", "核心正向场景"],
        ["LG-XM-MD-TC-010", "APP_REGISTER-注册/首登", "P0", "验证同一设备首登后再次登录不重复上报APP_REGISTER", "同一设备已经完成过首次登录并产生APP_REGISTER；当前已退出登录", "1.再次登录同一账号\n2.退出后再登录其他账号\n3.分别查看埋点数据", "设备首次登录=否", "后续重复登录仅上报APP_ACTIVE；APP_REGISTER在该设备维度不再上报；不会因更换账号重复首登", "PRD说明：仅限设备第一次登录行为"],
        ["LG-XM-MD-TC-011", "APP_REGISTER-注册/首登", "P1", "验证首次打开但未登录不应上报APP_REGISTER", "小米广告渠道首次安装；设备未登录过", "1.首次打开App\n2.停留在登录页或游客浏览（如支持）\n3.不完成登录\n4.查看埋点数据", "登录状态=未登录", "未完成登录不产生APP_REGISTER；不产生APP_ACTIVE；直到登录成功才按规则上报", "上报时机边界"],
        ["LG-XM-MD-TC-012", "APP_REGISTER-注册/首登", "P1", "验证首登接口重试或页面刷新不会重复上报APP_REGISTER", "小米广告渠道首次登录场景；可模拟登录成功回调重试、弱网重发或页面刷新", "1.构造弱网或接口重试\n2.完成首次登录\n3.查看同一设备同一首登链路的埋点条数", "设备首次登录=是\n网络=弱网/重试", "同一设备首登仅产生1条APP_REGISTER；不因接口重试、页面刷新、客户端重放导致重复上报", "防重复上报"],
        ["LG-XM-MD-TC-013", "APP_REGISTER-注册/首登", "P1", "验证非小米广告渠道首登不向小米上报APP_REGISTER", "全新设备通过非小米渠道安装App并首次登录", "1.从非小米广告渠道安装App\n2.首次打开并登录\n3.查看小米埋点数据", "渠道=非小米广告渠道\n设备首次登录=是", "不向小米上报APP_REGISTER；不得错误映射为小米广告首登", "渠道边界"],

        ["LG-XM-MD-TC-014", "APP_PAY-付费", "P0", "验证通过付费广告下载App后支付成功上报APP_PAY", "用户通过付费广告下载App；账号可购买会员/时长/订阅等商品；支付环境可用", "1.通过付费广告下载安装App\n2.登录并进入付费页\n3.选择商品完成支付\n4.查看支付成功后的埋点", "下载来源=付费广告\n支付结果=成功", "App内付费行为支付成功后上报APP_PAY；对应羚羊payment_successful；订单号、金额、币种、商品信息（如有）与订单一致", "核心正向场景"],
        ["LG-XM-MD-TC-015", "APP_PAY-付费", "P0", "验证付费广告下载后多次支付均上报APP_PAY", "同一付费广告来源用户可连续完成多笔支付", "1.完成第一笔支付并记录埋点\n2.再次购买其他商品或续费\n3.查看第二笔支付埋点", "支付次数=多次成功支付", "每次支付成功均上报APP_PAY；不同订单分别对应不同上报记录；不因已支付过而漏报", "PRD说明：app里面的付费行为都上报"],
        ["LG-XM-MD-TC-016", "APP_PAY-付费", "P1", "验证支付失败或取消不应上报APP_PAY", "付费广告来源用户；可取消支付或模拟支付失败", "1.发起订单\n2.在收银台取消支付或构造失败\n3.查看埋点数据", "支付结果=取消/失败", "未支付成功不产生APP_PAY；不产生羚羊payment_successful；订单状态与埋点一致", "负向场景"],
        ["LG-XM-MD-TC-017", "APP_PAY-付费", "P1", "验证非付费广告下载用户支付不向小米上报APP_PAY", "用户通过自然量或非付费广告安装App；可完成支付", "1.非付费广告渠道安装App\n2.登录并完成支付\n3.查看小米埋点数据", "下载来源=非付费广告\n支付结果=成功", "不向小米上报APP_PAY；内部支付成功事件如上报，应不被错误归因为小米付费广告", "来源边界"],
        ["LG-XM-MD-TC-018", "APP_PAY-付费", "P1", "验证支付成功回调重复不会导致APP_PAY重复上报", "付费广告来源用户；可模拟支付回调重试或客户端重复拉取订单结果", "1.完成一笔支付\n2.模拟支付成功回调重复通知或客户端重复确认订单\n3.查看APP_PAY上报条数", "同一订单号重复回调", "同一支付订单只上报一次APP_PAY；不会因回调重试产生重复付费埋点", "幂等性"],

        ["LG-XM-MD-TC-019", "归因与数据质量", "P0", "验证小米广告归因参数在三类事件中一致", "已分别触发APP_ACTIVE、APP_REGISTER、APP_PAY；可导出埋点明细", "1.导出三类事件数据\n2.核对渠道、设备、账号、时间戳、事件映射字段\n3.与广告归因记录对比", "事件=APP_ACTIVE/APP_REGISTER/APP_PAY", "三类事件均能正确关联小米广告归因；设备ID/用户ID/渠道标识一致且可追踪；事件时间与实际操作时间误差在可接受范围", "端到端校验"],
        ["LG-XM-MD-TC-020", "归因与数据质量", "P1", "验证卸载重装后设备首登规则符合设备维度定义", "同一设备曾通过小米广告渠道登录并上报APP_REGISTER；可卸载重装App", "1.卸载App\n2.通过小米广告渠道重新安装\n3.再次登录\n4.查看APP_ACTIVE与APP_REGISTER", "设备=同一设备\n操作=卸载重装后登录", "如服务端仍识别为同一设备且已有首登记录，则仅上报APP_ACTIVE不重复APP_REGISTER；如产品定义允许重置，应以需求确认口径为准并记录结果", "设备维度边界，若口径未明确需产品确认"],
        ["LG-XM-MD-TC-021", "归因与数据质量", "P1", "验证跨账号登录时注册事件仍按设备首次登录控制", "小米广告渠道安装；设备A已用账号1完成首登", "1.退出账号1\n2.使用账号2登录同一设备\n3.查看埋点", "设备=同一设备\n账号=不同账号", "账号2登录成功上报APP_ACTIVE；不再上报APP_REGISTER；首登判定不应从账号维度误判", "设备维度首登"],
        ["LG-XM-MD-TC-022", "归因与数据质量", "P1", "验证同一账号在新设备小米广告首登上报APP_REGISTER", "账号已在设备A登录过；设备B为小米广告渠道首次安装且设备B无登录记录", "1.在设备B通过小米广告安装App\n2.使用已存在账号登录\n3.查看埋点", "设备=新设备\n账号=老账号", "设备B第一次登录成功应上报APP_REGISTER和APP_ACTIVE；说明注册事件口径为设备首次登录而非账号注册", "设备维度首登"],
        ["LG-XM-MD-TC-023", "归因与数据质量", "P2", "验证弱网/离线恢复后三类事件上报策略符合预期", "可模拟弱网、断网和网络恢复；埋点SDK支持缓存或失败处理", "1.在弱网/断网下完成登录或支付\n2.恢复网络\n3.查看事件是否补发及条数", "网络=弱网/离线恢复", "业务成功的登录/支付事件按SDK策略补发或记录失败；不出现重复大量补发；事件字段完整未损坏", "稳定性"],
        ["LG-XM-MD-TC-024", "归因与数据质量", "P2", "验证埋点平台接收数据与客户端/服务端日志一致", "客户端日志、服务端日志、小米/羚羊平台查询权限可用", "1.执行一轮首登+重复登录+支付链路\n2.记录客户端与服务端日志\n3.在平台按设备/账号/时间查询\n4.对比事件条数和字段", "测试账号/设备/订单号", "平台侧可查询到对应事件；APP_REGISTER仅1次，APP_ACTIVE按登录成功次数，APP_PAY按支付成功订单数；字段与日志一致或符合服务端处理规则", "验收闭环"],
    ]

    smoke_headers = ["冒烟编号", "关联用例编号", "冒烟场景", "验证重点"]
    smokes = [
        ["SMK-LG-XM-001", "LG-XM-MD-TC-004", "小米渠道登录激活", "登录成功上报APP_ACTIVE且映射client_install"],
        ["SMK-LG-XM-002", "LG-XM-MD-TC-009", "小米渠道设备首登", "首次登录上报APP_REGISTER且映射login_on"],
        ["SMK-LG-XM-003", "LG-XM-MD-TC-010", "重复登录不重复注册", "后续登录仅上报APP_ACTIVE，不再上报APP_REGISTER"],
        ["SMK-LG-XM-004", "LG-XM-MD-TC-014", "付费广告支付成功", "支付成功上报APP_PAY且映射payment_successful"],
        ["SMK-LG-XM-005", "LG-XM-MD-TC-016", "支付失败/取消", "支付未成功不产生APP_PAY"],
        ["SMK-LG-XM-006", "LG-XM-MD-TC-019", "归因与字段", "渠道、设备、用户、事件映射字段正确一致"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "埋点测试用例"
    ws.append(headers)
    for row in cases:
        ws.append(row)

    smoke = wb.create_sheet("冒烟测试用例")
    smoke.append(smoke_headers)
    for row in smokes:
        smoke.append(row)

    apply_style(ws)
    apply_style(smoke)
    autosize_columns(ws, [18, 20, 10, 46, 38, 52, 32, 58, 22])
    autosize_columns(smoke, [16, 20, 30, 54])

    wb.save(output_path)
    return output_path, len(cases), len(smokes)


if __name__ == "__main__":
    print(build_workbook())