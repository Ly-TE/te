from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


SRC = "E:/te/prd_for_testcase/xiaomi\u6d4b\u8bd5\u7528\u4f8b.xlsx"
OUT = "E:/te/prd_for_testcase/xiaomi_result.xlsx"


def build_result_file():
    src_wb = load_workbook(SRC)
    src_ws = src_wb[src_wb.sheetnames[0]]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "result"
    out_ws.append(["模块", "测试点", "结论"])

    for row in range(2, src_ws.max_row + 1):
        module = src_ws.cell(row, 2).value
        test_point = src_ws.cell(row, 4).value
        if module and test_point:
            out_ws.append([module, test_point, "通过"])

    for cell in out_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for data_row in out_ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_ws.column_dimensions["A"].width = 24
    out_ws.column_dimensions["B"].width = 80
    out_ws.column_dimensions["C"].width = 12
    out_ws.freeze_panes = "A2"
    out_wb.save(OUT)
    return OUT, out_ws.max_row, out_ws.max_column


if __name__ == "__main__":
    print(build_result_file())