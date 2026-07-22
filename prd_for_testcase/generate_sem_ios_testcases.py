from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def autosize_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def build_workbook(output_path):
    headers = ["用例编号", "模块", "优先级", "用例标题", "前置条件", "测试步骤", "测试数据", "预期结果", "备注"]
    cases = [
        ["SEM-IOS-TC-001", "首页套餐", "P0", "验证 iOS 用户打开首页展示 2 个微信小店商品替换原充值套餐", "1.iOS 设备登录 SEM 小程序\n2.进入首页", "无特殊测试数据", "首页 2 个套餐区域展示为对应 2 个微信小店商品，不再显示原虚拟支付充值套餐；点击商品可进入对应购买/详情链路", "适合自动化"],
        ["SEM-IOS-TC-002", "首页套餐", "P1", "验证非 iOS 用户打开首页仍展示原充值套餐", "1.安卓设备登录 SEM 小程序\n2.进入首页", "无特殊测试数据", "非 iOS 端首页套餐展示不受影响，仍按原逻辑展示", ""],
        ["SEM-IOS-TC-003", "大学生页面", "P0", "验证 iOS 用户未认证时大学生页不展示套餐", "1.iOS 用户未实名认证\n2.访问 /Activity/index/studentNew2", "无特殊测试数据", "认证前页面不展示任何套餐入口", "适合自动化"],
        ["SEM-IOS-TC-004", "大学生页面", "P0", "验证 iOS 用户实名认证成功后展示 168-9000 小时小店商品", "1.iOS 用户完成实名认证\n2.进入 /Activity/index/studentNew2", "168-9000 小时商品", "认证成功后页面展示微信小店商品，且商品信息与配置一致", ""],
        ["SEM-IOS-TC-005", "大学生页面", "P1", "验证 iOS 用户认证成功超过半小时后再次进入隐藏套餐", "1.iOS 用户完成实名认证\n2.等待超过 30 分钟\n3.再次进入 /Activity/index/studentNew2", "超过 30 分钟", "再次进入页面时套餐重新隐藏，不再展示已认证后的商品", "需求边界，需确认计时口径"],
        ["SEM-IOS-TC-006", "底部充值页", "P0", "验证 iOS 用户看不到底部充值页 Tab 入口", "1.iOS 设备登录 SEM 小程序\n2.查看底部导航", "无特殊测试数据", "iOS 端底部充值页 Tab 不展示", "适合自动化"],
        ["SEM-IOS-TC-007", "底部充值页", "P0", "验证 iOS 用户无法通过直达进入底部充值页", "1.iOS 设备登录 SEM 小程序\n2.尝试通过已知链接/入口进入充值页", "无特殊测试数据", "iOS 端进入后相关页面不可见或无内容，不影响其他页面使用", ""],
        ["SEM-IOS-TC-008", "底部充值页", "P1", "验证支付宝小程序底部充值页仍可正常展示", "1.支付宝小程序登录\n2.查看底部充值入口", "无特殊测试数据", "支付宝端充值页入口和页面展示不受影响", ""],
        ["SEM-IOS-TC-009", "活动模板页", "P0", "验证活动模板页充值套餐替换为微信小店商品", "1.iOS 设备打开 /Activity/template/index?id=27653\n2.查看充值区域", "无特殊测试数据", "页面内充值套餐替换为微信小店商品，不再展示原虚拟支付套餐", "适合自动化"],
        ["SEM-IOS-TC-010", "活动模板页", "P0", "验证活动模板页抽奖入口去除", "1.iOS 设备打开 /Activity/template/index?id=27653\n2.查看活动操作区", "无特殊测试数据", "抽奖相关入口、按钮、弹窗不展示或不可操作", ""],
        ["SEM-IOS-TC-011", "活动模板页", "P0", "验证另一活动模板页充值套餐替换为微信小店商品", "1.iOS 设备打开 /Activity/template/index_1?id=41374\n2.查看充值区域", "无特殊测试数据", "页面内充值套餐替换为微信小店商品", "适合自动化"],
        ["SEM-IOS-TC-012", "活动模板页", "P0", "验证另一活动模板页抽奖入口去除", "1.iOS 设备打开 /Activity/template/index_1?id=41374\n2.查看活动操作区", "无特殊测试数据", "抽奖相关入口、按钮、弹窗不展示或不可操作", ""],
        ["SEM-IOS-TC-013", "我的页", "P0", "验证 iOS 用户我的页隐藏我的订单和优惠券", "1.iOS 设备进入我的页\n2.查看功能列表", "无特殊测试数据", "iOS 端不展示“我的订单”和“优惠券”入口", "适合自动化"],
        ["SEM-IOS-TC-014", "我的页", "P1", "验证非 iOS 用户我的页仍展示我的订单和优惠券", "1.安卓/其他端进入我的页\n2.查看功能列表", "无特殊测试数据", "非 iOS 端入口展示不受影响", ""],
        ["SEM-IOS-TC-015", "首页时长展示页", "P0", "验证首页时长展示页面改为小店入口", "1.iOS 设备进入首页\n2.点击时长展示区域", "无特殊测试数据", "点击后进入微信小店入口或对应商品页，不再进入原时长展示页", "适合自动化"],
        ["SEM-IOS-TC-016", "首页时长展示页", "P1", "验证非 iOS 端时长展示页面保持原逻辑", "1.安卓设备进入首页\n2.点击时长展示区域", "无特殊测试数据", "非 iOS 端仍按原时长展示页面逻辑处理", ""],
        ["SEM-IOS-TC-017", "Banner 配置", "P0", "验证配置关闭时 iOS 点击 banner 无响应", "1.将配置开关关闭\n2.iOS 设备点击首页 banner", "配置关闭", "点击 banner 后无页面跳转、无弹窗、无购买链路触发", "高风险配置项"],
        ["SEM-IOS-TC-018", "Banner 配置", "P1", "验证配置开启时 iOS 点击 banner 恢复正常响应", "1.将配置开关开启\n2.iOS 设备点击首页 banner", "配置开启", "banner 按配置规则正常跳转或展示", ""],
        ["SEM-IOS-TC-019", "实名认证状态", "P0", "验证 iOS 未认证状态进入小程序不展示套餐", "1.iOS 用户未实名认证\n2.进入首页、大学生页、活动页", "无特殊测试数据", "未认证态各目标页面不展示套餐入口或商品信息", "适合自动化"],
        ["SEM-IOS-TC-020", "实名认证状态", "P0", "验证 iOS 认证成功后展示小店商品且状态一致", "1.iOS 用户完成实名认证\n2.进入首页和大学生页\n3.刷新并切换页面", "无特殊测试数据", "认证成功后对应页面展示小店商品，前后端状态一致且页面切换后展示一致", ""],
        ["SEM-IOS-TC-021", "边界场景", "P1", "验证 iOS 认证成功后未满半小时仍展示套餐", "1.iOS 用户完成实名认证\n2.在 30 分钟内再次进入相关页面", "未满 30 分钟", "30 分钟内仍展示认证后的商品，不应提前隐藏", "需求边界，需确认计时口径"],
        ["SEM-IOS-TC-022", "兼容性", "P2", "验证 iOS 端页面改造后其他入口不受影响", "1.iOS 设备浏览首页、我的页、活动页\n2.检查其他非改造入口", "无特殊测试数据", "其他未提及入口可正常访问和使用", ""],
    ]

    smoke_headers = ["冒烟编号", "关联用例编号", "冒烟场景", "验证重点"]
    smokes = [
        ["SMK-001", "SEM-IOS-TC-001", "iOS 首页套餐替换", "首页 2 个套餐展示为微信小店商品"],
        ["SMK-002", "SEM-IOS-TC-003", "iOS 大学生页未认证隐藏套餐", "未认证态无套餐展示"],
        ["SMK-003", "SEM-IOS-TC-004", "iOS 大学生页认证后展示商品", "认证后展示 168-9000 小时商品"],
        ["SMK-004", "SEM-IOS-TC-006", "iOS 隐藏底部充值入口", "底部充值 Tab 不可见"],
        ["SMK-005", "SEM-IOS-TC-013", "iOS 我的页隐藏订单和优惠券", "指定入口不展示"],
        ["SMK-006", "SEM-IOS-TC-017", "banner 配置关闭", "点击 banner 无响应"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(headers)
    for row in cases:
        ws.append(row)

    smoke = wb.create_sheet("冒烟测试用例")
    smoke.append(smoke_headers)
    for row in smokes:
        smoke.append(row)

    for sheet in (ws, smoke):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"

    autosize_columns(ws, [14, 18, 10, 32, 30, 42, 24, 42, 20])
    autosize_columns(smoke, [14, 18, 28, 24])
    wb.save(output_path)


if __name__ == "__main__":
    build_workbook(r"prd_for_testcase\sem小程序ios相关页面修改测试用例.xlsx")